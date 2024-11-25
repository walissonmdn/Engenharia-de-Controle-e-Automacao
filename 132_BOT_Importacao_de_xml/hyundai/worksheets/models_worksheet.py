from openpyxl import load_workbook
from hyundai.constants import *
from hyundai.worksheets.excel import Excel

# Título da planilha.
WORKSHEET_TITLE = "Modelos"

# Colunas da planilha.
MODELO_COLUMN = "B"
OCN_COLUMN = "C"
ANO_COLUMN = "D"
CODIGO_DO_MODELO_COLUMN = "F"


class ModelsWorksheet(Excel):
    """Classe para interação com a planiha de relação de modelos."""

    def __init__(self):
        """Inicialização."""
        super().__init__(MAPPING_WORKBOOK_PATH, WORKSHEET_TITLE)

    def get_model_code(self, modelo, codigo_do_produto, ano_fabricacao, ano_modelo):
        """Compara a descrição do modelo do xml com a descrição do modelo na planilha de relação de
        modelos e busca o código do modelo para preencher no dealer.
        """
        # Ano de fabricação e ano do modelo no xml.
        year_in_the_xml = f"{ano_fabricacao[-2:]}/{ano_modelo[-2:]}" # Formato: "yy/yy".

        # Abre a planilha.
        wb = load_workbook(MAPPING_WORKBOOK_PATH)
        ws = wb[WORKSHEET_TITLE]

        # Inicialização de variável.
        model_code = None

        # Busca o código do modelo com base na descrição do veículo e no ano.
        for row in range(2, self.get_last_row() + 1):
            # Armazena o modelo e o ano em variáveis para realizar a comparação.
            model_in_the_worksheet = ws[MODELO_COLUMN + str(row)].value
            year_in_the_worksheet = ws[ANO_COLUMN + str(row)].value

            # Se o campo do modelo não estiver em branco, compara os valores armazenados com os
            # valores do xml.
            if model_in_the_worksheet != None:
                if ((modelo.strip().lower() == model_in_the_worksheet.strip().lower()) and
                    (year_in_the_xml.strip() == year_in_the_worksheet.strip())
                    ):
                    model_code = str(ws[CODIGO_DO_MODELO_COLUMN + str(row)].value)
                    break

        # Caso o código do modelo não tenha sido encontrado, tenta identificar o código do modelo 
        # com base no OCN e no ano.
        if model_code == None:
            for row in range(2, self.get_last_row() + 1):
                # Amarzena o OCN e o ano em variáveis.
                ocn_in_the_worksheet = ws[OCN_COLUMN + str(row)].value
                year_in_the_worksheet = ws[ANO_COLUMN + str(row)].value
                
                # Se o campo de OCN não estiver em branco, compara com o código do produto (o OCN
                # está contido no código), e compara o ano de fabricação e o ano do modelo.
                if ocn_in_the_worksheet != None:
                    if ((ocn_in_the_worksheet.strip().lower() in codigo_do_produto.strip().lower()) and
                        (year_in_the_xml.strip() == year_in_the_worksheet.strip())
                        ):
                        model_code = str(ws[CODIGO_DO_MODELO_COLUMN + str(row)].value)
                        break

        # Fecha e retorna o código do modelo.
        wb.close()
        return model_code
    

