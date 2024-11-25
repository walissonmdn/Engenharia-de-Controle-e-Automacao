from hyundai.constants import *
from hyundai.worksheets.excel import Excel
from openpyxl import load_workbook

# Título.
WORKSHEET_TITLE = "HYUNDAI"

# Colunas.
NOME_DO_ARQUIVO_COLUMN = "A"
DATA_DE_EMISSAO_COLUMN = "B"
NUMERO_COLUMN = "C"
CNPJ_COLUMN = "D"
VALOR_COLUMN = "E"
CHASSI_COLUMN = "F"
MODELO_COLUMN = "G"
COR_EXTERNA_COLUMN = "H"
INFORMACOES_DO_PEDIDO_COLUMN = "I"
STATUS_COLUMN = "J"
OBSERVACAO_COLUMN = "K"


class ResultsWorksheet(Excel):
    """Manipula a planilha de resultados."""

    def __init__(self):
        """Inicialização."""
        super().__init__(RESULTS_WORKBOOK_PATH, WORKSHEET_TITLE)
    
    def get_xml_row(self, filename_in_datapool):
        """Identifica se o arquivo a ser importado já se encontra na planilha e retorna a linha. Se 
        não encontrar, retorna a primeira linha disponível na planilha para preenchimento.
        """
        # Abre a planilha.
        wb = load_workbook(RESULTS_WORKBOOK_PATH)
        ws = wb[WORKSHEET_TITLE]

        # Inicialização de variável.
        row = ws.max_row + 1

        # Verifica se já existe dados sobre o arquivo na planilha e se houver, armazena a linha.
        for filename_cell in ws[f"{NOME_DO_ARQUIVO_COLUMN}:{NOME_DO_ARQUIVO_COLUMN}"]:
            if filename_cell.value.strip().lower() == filename_in_datapool.strip().lower():
                row = filename_cell.row
                break

        # Fecha a planilha e retorna o número da linha.
        wb.close()
        return row
    
    def fill_in(self, datapool_item, row):
        """Preenche os campos da planilha com os dados do Datapool."""
        # Abertura da planilha.
        wb = load_workbook(RESULTS_WORKBOOK_PATH)
        ws = wb[WORKSHEET_TITLE]

        # Preenchimento dos campos.
        ws[NOME_DO_ARQUIVO_COLUMN + str(row)].value = datapool_item["nome_do_arquivo"]
        ws[DATA_DE_EMISSAO_COLUMN + str(row)].value = datapool_item["data_de_emissao"]
        ws[NUMERO_COLUMN + str(row)].value = datapool_item["nota_fiscal"]
        ws[CNPJ_COLUMN + str(row)].value = datapool_item["cnpj"]
        ws[VALOR_COLUMN + str(row)].value = datapool_item["valor"]
        ws[CHASSI_COLUMN + str(row)].value = datapool_item["chassi"]
        ws[MODELO_COLUMN + str(row)].value = datapool_item["modelo_do_veiculo"]
        ws[COR_EXTERNA_COLUMN + str(row)].value = datapool_item["cor_externa"]
        ws[INFORMACOES_DO_PEDIDO_COLUMN + str(row)].value = datapool_item["informacoes_do_pedido"]

        # Salva e fecha a planilha.
        wb.save(RESULTS_WORKBOOK_PATH)
        wb.close()