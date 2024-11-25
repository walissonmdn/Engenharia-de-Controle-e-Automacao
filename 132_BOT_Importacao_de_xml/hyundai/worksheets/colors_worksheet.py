from openpyxl import load_workbook
from hyundai.constants import *
from hyundai.worksheets.excel import Excel

# Título da planilha.
WORKSHEET_TITLE = "Cores"

# Colunas da planilha.
COLOR_COLUMN = "B"
COLOR_CODE_COLUMN = "C"


class ColorsWorksheet(Excel):
    """Classe para interação com a planiha de relação de modelos."""

    def __init__(self):
        """Inicialização."""
        super().__init__(MAPPING_WORKBOOK_PATH, WORKSHEET_TITLE)

    def get_color(self, color, color_code):
        """"""
        wb = load_workbook(MAPPING_WORKBOOK_PATH)
        ws = wb[WORKSHEET_TITLE]

        # Busca o código do modelo com base na descrição do veículo e no ano.
        color_in_dealer = None
        color_code_in_dealer = None
        for row in range(2, ws.max_row + 1):
            color_in_worksheet = ws[COLOR_COLUMN + str(row)].value
            color_code_in_worksheet = ws[COLOR_CODE_COLUMN + str(row)].value

            print(color)
            print(color_in_worksheet)
            print(color_code)
            print(color_code_in_worksheet)
            if (color_in_worksheet == None) or (color_code_in_worksheet == None):
                continue

            if ((color.strip().lower() in color_in_worksheet.strip().lower()) and
                color_code_in_worksheet.strip().lower() == color_code.strip().lower()
                ):
                color_in_dealer = color_in_worksheet
                color_code_in_dealer = color_code_in_worksheet
                break

        wb.close()
        return color_in_dealer, color_code_in_dealer
    

