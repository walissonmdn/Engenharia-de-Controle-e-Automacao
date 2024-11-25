from openpyxl import load_workbook
from hyundai.constants import *


class Excel:
    """Classe para manipulação de planilhas do excel (.xlsx)."""

    def __init__(self, worksheet_path, worksheet_title):
        self.worksheet_path = worksheet_path
        self.worksheet_title = worksheet_title

    def delete_row(self, row):
        """Deleta da planilha a linha passada como parâmetro."""
        wb = load_workbook(self.worksheet_path)
        ws = wb[self.worksheet_title]

        ws.delete_rows(row)

        wb.save(self.worksheet_path)
        wb.close()
    
    def get(self, column, row):
        """Busca um dado."""
        wb = load_workbook(self.worksheet_path)
        ws = wb[self.worksheet_title]

        value = ws[str(column) + str(row)].value

        wb.close()
        return value

    def get_last_row(self):
        """Retorna o número da última linha."""
        wb = load_workbook(self.worksheet_path)
        ws = wb[self.worksheet_title]

        last_row = ws.max_row

        wb.close()
        return last_row
    
    def write(self, column, row, value):
        """Escreve em uma célula."""
        wb = load_workbook(self.worksheet_path)
        ws = wb[self.worksheet_title]

        ws[str(column) + str(row)].value = value

        wb.save(self.worksheet_path)
        wb.close()