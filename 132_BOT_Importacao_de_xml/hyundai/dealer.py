from hyundai.constants import *
from hyundai.worksheets.colors_worksheet import ColorsWorksheet
from hyundai.worksheets.models_worksheet import ModelsWorksheet
from hyundai.worksheets.results_worksheet import ResultsWorksheet
from hyundai.worksheets.stores_worksheet import StoresWorksheet
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import hyundai.worksheets.results_worksheet as results_worksheet 
import time


class Dealer:
    """Classe para interação com o dealer."""
    
    def __init__(self, selenium, maestro, log) :
        """Inicialização de variáveis."""
        self.selenium = selenium
        self.maestro = maestro
        self.log = log

        self.results_worksheet = ResultsWorksheet()
        self.stores_worksheet = StoresWorksheet()
        self.models_worksheet = ModelsWorksheet()
        self.colors_worksheet = ColorsWorksheet()

        # As variáveis abaixo serão alteradas durante a execução do programa.
        self.worksheet_current_row = None
        self.datapool_item = None
        self.xml_inserted = None

    def login(self, try_again=False):
        """Realiza login no Dealer."""
        self.selenium.get_page("https://dealer.gruposaga.com.br/")
        self.selenium.wait_for_element_to_appear("input#vUSUARIO_IDENTIFICADORALTERNATIVO")
        time.sleep(1)
        if try_again == False:
            self.selenium.fill("input#vUSUARIO_IDENTIFICADORALTERNATIVO", 
                               self.maestro.get_credential("dealer", "soulbot_login"))
        self.selenium.fill("input#vUSUARIOSENHA_SENHA", 
                           self.maestro.get_credential("dealer", "soulbot_password"))
        self.selenium.click("input#IMAGE3")

        # Aguarda até a loja aparecer, pois isso signica que o login foi bem-sucedido.
        self.selenium.wait_for_element_to_appear("tr.x-toolbar-right-row > td:nth-child(4) > table.x-btn.x-btn-noicon")
        self.log.info(message="Login realizado com sucesso.")

    def change_store(self):
        """Método para realizar a busca na planilha de relação de lojas a loja a ser selecionada no
        dealer. É feita uma leitura das linhas da planilha cujo CNPJ é comparado com o CNPJ obtido 
        do xml. Caso o CNPJ seja encontrado, a marca e a loja são armazenadas para serem 
        selecionadas no dealer. Caso não seja encontrado, retorna uma exceção. 
        """
        """Busca a marca e a loja a serem selecionadas no Dealer, verifica se já estão selecionadas
        e se não estiverem, seleciona-as. Se a loja não estiver na planilha de relação de lojas ou 
        houver erro ao selecionar, lança exceção.
        """
        # Retorna o nome da marca e da loja que serão selecionadas no Dealer.
        make_and_store = self.stores_worksheet.get_make_and_store(self.datapool_item["cnpj"])
        make_to_be_selected = make_and_store[0]
        store_to_be_selected = make_and_store[1]

        # Verifica se foi possível encontrar a marca e a loja na planilha.
        if store_to_be_selected == None:
            self.results_worksheet.write(results_worksheet.STATUS_COLUMN, 
                                     self.worksheet_current_row, 
                                     "CNPJ não encontrado na planilha de relação de lojas para selecionar a loja no dealer.")
            raise Exception("CNPJ não encontrado na planilha de relação de lojas para selecionar a loja no dealer.")
        
        # Verifica se a loja já está selecionada.
        self.selenium.switch_to_default_content()
        selected_store_tr_element = self.selenium.find_elements("tr.x-toolbar-right-row")[-1]
        selected_store = selected_store_tr_element.find_element("button").text()
        if selected_store.lower().strip() == store_to_be_selected.strip().lower():
            self.log.info(message="A loja correta já está selecionada.")
            return

        # Inicizalização de variáveis.
        max_attempts = 4
        correct_store_selected = False
            
        for _ in range(max_attempts):
            # Clique para garantir que a lista de lojas será fechada para fazer uma nova tentativa.
            self.selenium.click("div#W5-timeout-menu")
            time.sleep(0.5)

            # Expande a lista de marcas.
            self.selenium.click("tr.x-toolbar-right-row > td:nth-child(4) > table.x-btn.x-btn-noicon")
            self.selenium.wait_for_element_to_appear("li:nth-child(1) > a.x-menu-item.x-menu-item-arrow.x-unselectable > span")
            
            # Procura pela marca.
            makes = self.selenium.find_elements("li > a.x-menu-item.x-menu-item-arrow.x-unselectable > span")
            for make in makes: 
                if make.text().strip().lower() == make_to_be_selected.strip().lower():
                    make.hover()
                    break
            time.sleep(0.5) 

            # Procura pela loja.
            stores_div = self.selenium.find_elements("div.x-menu.x-menu-floating")[-1]
            stores = stores_div.find_elements("span.x-menu-item-text")
            for store in stores:
                print(f'"{store.text()}"')
                print(f'"{store_to_be_selected}"')
                if store.text().strip().lower() == store_to_be_selected.strip().lower():
                    store.click()
                    break  

            # Verifica a loja selecionada..
            selected_store = self.selenium.find_element("tr.x-toolbar-right-row > td:nth-child(4) > table.x-btn.x-btn-noicon > tbody > tr:nth-child(2) > td:nth-child(2) > em > button.x-btn-text").text()
            if selected_store.lower().strip() == store_to_be_selected.strip().lower():
                correct_store_selected = True
                break
        
        # Armazena no log e na planilha se a loja foi selecionada corretamente ou não.
        if correct_store_selected:
            self.log.info(message="A loja foi selecionada no Dealer.")
        else:
            self.results_worksheet.write(results_worksheet.STATUS_COLUMN, 
                                     self.worksheet_current_row, 
                                     "Não foi possível selecionar a loja no Dealer.")
            raise Exception("Não foi possível selecionar a loja no Dealer.")

    def import_xml(self, xml_path):
        """Navega até a página de importação do xml e realiza a importação."""
        # Navega até a página de importação do xml.
        self.selenium.click("table[id*= 'Integração']") 
        self.selenium.hover("a[id*= 'XML-Importação']")
        self.selenium.click("a[id*= 'NotaFiscaldeVeículo']")
        self.selenium.switch_to_frame("div.x-window-body > iframe")
        
        # A variável self.xml_inserted serve como indicativo a respeito da importação do xml. Caso 
        # o xml tenha sido importado corretamente e o processo tenha retornado ao início em 
        # decorrência de algum problema de interação com o dealer, o processo será feito do início,
        # mas sem importar o xml novamente.3
        if self.xml_inserted == False: 
            # Botão para direcionar à pagina de importação.
            self.selenium.click("input#IMAGE1")
            
            # Importação.
            self.selenium.click("input#IMAGE2")
            self.selenium.switch_to_frame("iframe#gxp0_ifrm")
            self.selenium.wait_for_element_to_appear("input#uploadfiles")
            time.sleep(0.5)
            self.selenium.upload_file("input#uploadfiles", xml_path)
            self.selenium.switch_to_default_content()
            self.selenium.switch_to_frame("div.x-window-body > iframe")
            self.selenium.click("input#TRN_ENTER")
            self.selenium.wait_for_element_to_appear("span#TEXTBLOCKDOWNLOAD > a > text")

            # Verifica se a importação ocorreu com sucesso.
            message = self.selenium.find_element("span#TEXTBLOCKDOWNLOAD > a > text").text()
            if "1 de 1" in message:
                self.xml_inserted = True
                self.log.info(message="Nota importada com sucesso.")
            else:
                self.results_worksheet.write(results_worksheet.STATUS_COLUMN, 
                                             str(self.worksheet_current_row), 
                                             "Não foi possível importar a nota.")
                raise Exception("Não foi possível importar a nota. Verifique a loja selecionada ou o xml.")

            # Volta para a página anterior.
            self.selenium.click("input#TRN_CANCEL")
        else:
            self.log.info(message="Nota já importada.")

    def fill_in_informations(self):
        """Verifica em qual linha da tabela do dealer se encontra a nota importada e preenche na 
        página de dados do veículo os dados obtidos do xml.
        """
        # Aguarda até que a primeira linha da tabela apareça.
        self.selenium.wait_for_element_to_appear("table#GridintxmlContainerTbl") 
        self.selenium.delete_element("table#GridintxmlContainerTbl")
        self.selenium.wait_for_element_to_disappear("table#GridintxmlContainerTbl")
        self.selenium.clear("input#vINTEGRACAOXMLNF_DATAEMISSAOINICIAL")
        self.selenium.wait_for_element_to_appear("table#GridintxmlContainerTbl") 

        # Encontra a linha com a nota importada e clica no ícone para editar as informações.
        row_found = False # Inicialização de variável.
        table = self.selenium.find_elements("table#GridintxmlContainerTbl > tbody > tr")
        for table_row in table:
            document_number = table_row.find_element("td:nth-child(4) > span").text()
            if document_number == self.datapool_item["nota_fiscal"]:
                row_found = True
                table_row.find_element("td:nth-child(9) > a > img").click()
                break

        # Se a nota não tiver aparecido na tabela, informar na planilha e prosseguir para a próxima.
        if row_found == False:
            self.results_worksheet.write(results_worksheet.STATUS_COLUMN,
                                         self.worksheet_current_row,
                                         "Linha com a nota inserida não encontrada na tabela do dealer após importação do xml.")
            raise Exception("Linha com a nota inserida não encontrada na tabela do dealer após importação do xml.")
        
        # Prossegue para a página "Dados Veículo".
        self.selenium.click("input#vUPDATE_0001")

        # Cor a ser selecionada como opcional e em cor externa.
        color_in_worksheet = self.colors_worksheet.get_color(self.datapool_item["cor_externa"], 
                                                             self.datapool_item["codigo_da_cor_externa"]) # Retorna uma tupla.
        
        # Se a cor for encontrada na planilha, armazena a cor e o código. Se não, utiliza a cor e 
        # código do 
        if color_in_worksheet[0] != None:
            color_in_dealer = color_in_worksheet[0]
            color_code_in_dealer = color_in_worksheet[1]
        else:
            color_in_dealer = self.datapool_item["cor_externa"]
            color_code_in_dealer = self.datapool_item["codigo_da_cor_externa"]

        # Opcional (não será executado para as lojas de MT).
        if ((self.datapool_item["cnpj"] != "22280413000290") and
            (self.datapool_item["cnpj"] != "22280413000100")
            ):
            self.selenium.click("span#TEXTBLOCK2")
            self.selenium.click("input#INCLUIR")
            self.selenium.switch_to_frame("iframe#gxp0_ifrm")
            
            self.selenium.wait_for_element_to_appear("table#GridselectedContainerTbl")
            time.sleep(0.5)
            try:
                self.selenium.find_element("input#vDELETE_0001")
            except:
                pass
            else:
                self.selenium.click("input#vDELETE_0001")
                time.sleep(0.5)
                self.selenium.wait_for_element_to_disappear("input#vDELETE_0001")

            self.selenium.delete_element("table#GridContainerTbl")
            self.selenium.script(f"document.getElementById('vDESCRICAO').value = '{color_in_dealer}'")
            self.selenium.script(f"document.getElementById('vSIGLA').value = '{color_code_in_dealer}'")
            time.sleep(0.5)

            self.selenium.wait_for_element_to_disappear("table#GridContainerTbl")
            self.selenium.click("#IMGREFRESH")
            self.selenium.wait_for_element_to_appear("table#GridContainerTbl")
            time.sleep(1)

            # Tenta buscar o texto em "Opcional" e em "Sigla" e se encontrar, compara com o texto
            # extraído do xml para selecionar no dealer.
            results_found = self.selenium.element_exists("span#span_vOPCIONAL_DESCRICAO_0001")

            if results_found:
                opcional = self.selenium.find_element("span#span_vOPCIONAL_DESCRICAO_0001").text()
                sigla = self.selenium.find_element("span#span_vOPCIONAL_SIGLA_0001").text()

                if (
                    (opcional.strip().lower() == color_in_dealer.strip().lower()) and
                    (sigla.strip().lower() == color_code_in_dealer.strip().lower())
                    ):
                    self.selenium.click("input#vSELECT_0001")
                    self.selenium.wait_for_element_to_appear("span#span_vOPCIONAL_DESCRICAOSELECTED_0001")
                    self.results_worksheet.write(results_worksheet.OBSERVACAO_COLUMN,
                                        self.worksheet_current_row,
                                        "")
                else:
                    self.results_worksheet.write(results_worksheet.OBSERVACAO_COLUMN,
                                        self.worksheet_current_row,
                                        "Opcional não encontrado.")
            else:
                self.results_worksheet.write(results_worksheet.OBSERVACAO_COLUMN,
                                        self.worksheet_current_row,
                                        "Opcional não encontrado.")                
            
            # Botão de confirmação.
            self.selenium.click("input#TRN_ENTER")

            # Navega até a aba "Geral".
            self.selenium.switch_to_default_content()
            self.selenium.switch_to_frame("div.x-window-body > iframe")
            self.selenium.click("span#TXBGERAL")
        
        # Página: "Dados Veículo";
        self.selenium.wait_for_element_to_appear("select#vMARCA_CODIGO")
        self.log.info(message="Inserindo os dados do veículo.")


        # Marca.
        self.selenium.select("//select[@id='vMARCA_CODIGO']", "HYUNDAI", By.XPATH)

        # Modelo.
        model_code = self.models_worksheet.get_model_code(self.datapool_item["modelo_do_veiculo"],
                                                          self.datapool_item["codigo_do_produto"],
                                                          self.datapool_item["ano_fabricacao"],
                                                          self.datapool_item["ano_modelo"])

        if model_code == None:
            self.log.error("Modelo não encontrado na planilha de relação de modelos.")
            self.results_worksheet.write(results_worksheet.STATUS_COLUMN,
                                     self.worksheet_current_row,
                                     "Modelo não encontrado na planilha de relação de modelos.")
            raise Exception("Modelo não encontrado na planilha de relação de modelos.")

        self.selenium.find_element("input#vMODELOVEICULO_CODIGO").clear()
        self.selenium.fill("input#vMODELOVEICULO_CODIGO", model_code + Keys.TAB)

        # Cor interna (Será sempre PRETO/CINZA para a Hyundai).
        self.selenium.select("//select[@id='vINTEGRACAO_CORCODINTERNA']", "PRETO/CINZA", By.XPATH)

        # Cor externa.
        print(f"cor externa: '{color_in_dealer.upper()}'")
        selection_result = self.selenium.select("//select[@id='vINTEGRACAO_CORCODEXTERNA']", 
                                             color_in_dealer.upper(), 
                                             By.XPATH, 
                                             wait_time=60, 
                                             return_exception=True)
        if selection_result == "Timeout.":
            self.results_worksheet.write(results_worksheet.STATUS_COLUMN,
                                     self.worksheet_current_row,
                                     "Cor externa não encontrada no Dealer.")
            raise Exception("Cor externa não encontrada no Dealer.")

        # Combustível.
        tipo_de_combustivel = None
        wb = load_workbook(FUEL_WORKBOOK_PATH)
        ws = wb["Combustível"]
        for fuel_code_cell in ws["A:A"]:
            if fuel_code_cell.value.zfill(2) == self.datapool_item["tipo_de_combustivel"]:
                tipo_de_combustivel = ws["B" + str(fuel_code_cell.row)].value
                self.selenium.select("//select[@id='vCOMBUSTIVEL_CODIGO']", 
                                     tipo_de_combustivel, 
                                     By.XPATH)
                break
        if tipo_de_combustivel == None:
            raise Exception("Código do combustível não encontrado.")
    
        # Procedência.
        self.selenium.select("//select[@id='vINTEGRACAOXMLNFVEICULO_PROCEDENCIACOD']", 
                             "0-PRODUTO NACIONAL", 
                             By.XPATH)
        
        # Clica no botão de confirmação.
        self.selenium.click("input#BTNCONFIRMAR")

        # Inicialização de variável.
        error_message = None
        
        # Verifica se houve erro.
        shown_element = self.selenium.wait_for_one_of_the_elements_to_appear("table[id^='TABLE'] > tbody > tr > td > div > span#gxErrorViewer > div",
                                                                             "input#BTNPROCESSAR")
        if shown_element == 1:
            error_message = self.selenium.find_element("table[id^='TABLE'] > tbody > tr > td > div > span#gxErrorViewer > div").text()
            self.results_worksheet.write(results_worksheet.STATUS_COLUMN,
                                 self.worksheet_current_row,
                                 f'Mensagem de erro: "{error_message}"')
        else:
            self.log.info(message="Dados preenchidos com sucesso.")
        
        return error_message

        
    def process_data(self):
        """Campos que aparecem para preenchimento ao clicar no botão "Processar"."""
        # Navegando até a página.
        self.selenium.click("input#BTNPROCESSAR", wait_time=10)

        self.log.info(message="Página de processamento de dados.")
        
        # Grupo Movimento.
        self.selenium.select("select#vNOTAFISCAL_GRUPOMOVIMENTO", "Compra")
        
        # Natureza Operação.
        natureza_operacao = self.selenium.find_element("span#span_vNOTAFISCAL_NATUREZAOPERACAODES").text()
        if ("compra de veiculo novo" in natureza_operacao.lower()) == False:
            self.selenium.click("input#NATOPE")
            self.selenium.switch_to_frame("iframe#gxp0_ifrm")
            self.selenium.select("select#vNOTAFISCAL_NATUREZAOPERACAOCOD", "COMPRA DE VEICULO NOVO")
            self.selenium.click("input#CONFIRMAR")
            self.selenium.switch_to_default_content()
            self.selenium.switch_to_frame("div.x-window-body > iframe")
        
        # Tipo Pessoa.
        self.selenium.select("select#vPESSOA_TIPOPESSOA", "Montadora")

        # Tipo Documento.
        tipo_documento = self.selenium.find_element("span#span_vNOTAFISCAL_TIPODOCUMENTODES").text()
        if ("nf-e - nota fiscal eletronica" in tipo_documento.lower()) == False:
            self.selenium.click("input#TIPODOC")
            self.selenium.switch_to_frame("iframe#gxp0_ifrm")
            self.selenium.select("select#vNOTAFISCAL_TIPODOCUMENTOCOD", "NF-E - NOTA FISCAL ELETRONICA")
            self.selenium.click("input#CONFIRMAR")
            self.selenium.switch_to_default_content()
            self.selenium.switch_to_frame("div.x-window-body > iframe")

        # Conta Gerencial.
        self.selenium.click("input#vNOTAFISCAL_CONTAGERENCIALCOD")
        conta_gerencial = self.selenium.find_element("input#vNOTAFISCAL_CONTAGERENCIALCOD")
        conta_gerencial.clear()
        conta_gerencial.send_keys("68")

        # Departamento
        self.selenium.select("select#vNOTAFISCAL_DEPARTAMENTOCOD", "VEICULOS NOVOS")

        # Estoque.
        self.selenium.select("select#vNOTAFISCALITEM_ESTOQUECOD", "VN - VEICULOS NOVOS")
        
        # Condição Pagamento.
        self.selenium.select("select#vNOTAFISCAL_CONDICAOPAGAMENTOCOD", 
                             f"VEICULOS NOVOS HYUNDAI")

        # Agente Cobrador.
        self.selenium.select("select#vAGENTECOBRADOR_CODIGO", "VEICULOS")

        # Tipo de Faturamento.
        self.selenium.select("select#vVEICULOTIPOFATURAMENTO_CODIGO", 
                             f"FLOOR PLAN - HYUNDAI")

        # Regra dos Tributos.
        considerar_regras_dos_tributos = self.selenium.find_element("input#vISUTILIZAREGRATRIBUTOICMSPISCOFINS")
        if considerar_regras_dos_tributos.is_selected() == True:
            considerar_regras_dos_tributos.click()

        # Finaliza o processo.
        self.selenium.click("input#IMGPROCESSAR")

        # Aguarda a mensagem de retorno e armazena.
        self.selenium.wait_for_element_to_appear("#TABLE1 > tbody > tr > td > div >span#gxErrorViewer > div")
        messages = self.selenium.find_elements("table#TABLE1 > tbody > tr > td > div >span#gxErrorViewer > div")

        # Formata a mensagem para que apareça uma mensagem por linha.
        result = ""
        for i in range(len(messages)):
            result += messages[i].text()
            if i != (len(messages) -1):
                result += "\n"
        self.log.info(message=f'Mensagem de resultado: "{result}"')
        
        # Armazena a mensagem na plinha.
        self.results_worksheet.write(results_worksheet.STATUS_COLUMN,
                                 self.worksheet_current_row,
                                 result)