from openpyxl import load_workbook
from kia.constants import *
from kia.worksheets.excel import Excel

# Título da planilha.
WORKSHEET_TITLE = "Modelos"

# Colunas da planilha.
DESCRICAO_COLUMN = "A"
CODIGO_COLUMN = "B"
CODIGO_DEALER_COLUMN = "D"


class ModelsWorksheet(Excel):
    """Classe para interação com a planiha de relação de modelos."""

    def __init__(self):
        """Inicialização."""
        super().__init__(MAPPING_WORKBOOK_PATH, WORKSHEET_TITLE)

    def get_model_code(self, description, code):
        """Compara a descrição do modelo do xml com a descrição do modelo na planilha de relação de
        modelos e busca o código do modelo para preencher no dealer.
        """
        # Abre a planilha de modelos.
        wb = load_workbook(MAPPING_WORKBOOK_PATH)
        ws = wb[WORKSHEET_TITLE]

        # Inicialização de variável.
        model_code = None 

        # Busca pelo código do modelo com base na descrição do veículo e no ano.
        for row in range(2, ws.max_row + 1):

            # Armazena a descrição e o código que estão na planilha.
            description_in_the_worksheet = ws[DESCRICAO_COLUMN + str(row)].value
            code_in_the_worksheet = ws[CODIGO_COLUMN + str(row)].value

            # Caso os campos não estejam em branco.
            if (description_in_the_worksheet != None) and (code_in_the_worksheet != None):
                # Compara a descrição e o código do modelo dos argumentos com aqueles da planliha.
                if ((description.strip().lower() == description_in_the_worksheet.strip().lower()) and
                    (code.strip() == code_in_the_worksheet.strip())
                    ):
                    model_code = str(ws[CODIGO_DEALER_COLUMN + str(row)].value)
                    break

        # Fecha a planilha e retorna o modelo.
        wb.close()
        return model_code