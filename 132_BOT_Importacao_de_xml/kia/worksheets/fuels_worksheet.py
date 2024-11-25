from kia.constants import *
from kia.worksheets.excel import Excel
from openpyxl import load_workbook

# Título.
WORKSHEET_TITLE = "Combustíveis"

# Colunas.
CODIGO_COLUMN = "A"
DESCRICAO_COLUMN = "B"


class FuelsWorksheet(Excel):
    """Manipula a planilha de combustível."""

    def __init__(self):
        """Inicialização."""
        super().__init__(STORES_WORKBOOK_PATH, WORKSHEET_TITLE)

    def get_fuel(self, fuel_code):
        """Busca na planilha o combustível através da comparação de códigos."""
        # Inicialização de variável
        fuel = None

        # Abre a planilha.
        wb = load_workbook(FUEL_WORKBOOK_PATH)
        ws = wb["Combustível"]
        # Percorre a coluna de código e compara cada um com o código do argumento.
        for row in range(2, ws.max_row + 1):
            fuel_code_in_worksheet = ws[CODIGO_COLUMN + str(row)].value
    
            if int(fuel_code_in_worksheet) == int(fuel_code):
                fuel = ws[DESCRICAO_COLUMN + str(row)].value
                break

        # Fecha a planilha e retorna o combustível.
        wb.close()
        return fuel
