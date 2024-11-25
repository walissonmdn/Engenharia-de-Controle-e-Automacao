from byd.constants import *
from byd.worksheets.excel import Excel
from openpyxl import load_workbook

# Título.
WORKSHEET_TITLE = "Empresas Dealer"

# Colunas.
PESSOA_NOME_COLUMN = "A"
CNPJ_COLUMN = "B"
EMPRESA_DEALER_COLUMN = "C"
MARCA_COLUMN = "D"


class StoresWorksheet(Excel):
    """Manipula a planilha de relação de lojas."""

    def __init__(self):
        """Inicialização."""
        super().__init__(STORES_WORKBOOK_PATH, WORKSHEET_TITLE)

    def get_make_and_store(self, cnpj_to_be_found):
        # Inicialização de variáveis.
        make = None
        store = None

        # Carrega o arquivo e a planilha.
        wb = load_workbook(STORES_WORKBOOK_PATH)
        ws = wb["Empresas Dealer"]

        # Procura pela marca e pela loja
        for cnpj_cell in ws[f"{CNPJ_COLUMN}:{CNPJ_COLUMN}"]:
            # Se a célula estiver em branco, passa para a próxima.
            if cnpj_cell.value == None:
                continue
            
            # Remove caracteres do cnpj retornado da planilha.
            cnpj_in_worksheet = cnpj_cell.value.replace(".", "").replace("/", "").replace("-", "")

            # Verifica se o CNPJ da linha da planilha é o mesmo CNPJ do argumento.
            if cnpj_in_worksheet == cnpj_to_be_found:
                make = ws[MARCA_COLUMN + str(cnpj_cell.row)].value
                store = ws[EMPRESA_DEALER_COLUMN + str(cnpj_cell.row)].value
                break

        # Fecha a conexão com a planilha e retorna os dados.
        wb.close()
        return make, store